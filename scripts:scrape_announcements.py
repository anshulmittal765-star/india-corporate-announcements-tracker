#!/usr/bin/env python3
"""
Indian Corporate Announcements Scraper
Fetches announcements from BSE India and extracts key highlights
"""

import requests
import pandas as pd
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
import json
import re
import os
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import PyPDF2
from io import BytesIO

# Configuration
BSE_ANNOUNCEMENTS_URL = "https://api.bseindia.com/BseIndiaAPI/api/AnnGetData/w"
BSE_PDF_BASE_URL = "https://www.bseindia.com/xml-data/corpfiling/AttachLive"
NSE_ANNOUNCEMENTS_URL = "https://www.nseindia.com/api/corporate-announcements"

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'en-US,en;q=0.9',
    'Referer': 'https://www.bseindia.com/',
    'Origin': 'https://www.bseindia.com'
}

# Categories to track
ANNOUNCEMENT_CATEGORIES = [
    'Board Meeting',
    'Financial Results',
    'Dividend',
    'AGM/EGM',
    'Acquisition',
    'Investment',
    'Fund Raising',
    'Merger/Demerger',
    'Change in Directors',
    'Corporate Action',
    'Investor Presentation',
    'Concall Transcript',
    'Order Win',
    'New Contract',
    'Expansion',
    'Capex',
    'Rating',
    'Others'
]

# Keywords for investment implications
POSITIVE_KEYWORDS = [
    'profit increase', 'profit up', 'revenue growth', 'dividend', 'bonus',
    'acquisition', 'expansion', 'new order', 'contract win', 'upgrade',
    'record', 'highest ever', 'beat estimates', 'outperform', 'growth',
    'investment', 'capex', 'expansion plan', 'new plant', 'capacity addition'
]

NEGATIVE_KEYWORDS = [
    'profit decline', 'profit down', 'revenue decline', 'loss', 'downgrade',
    'resign', 'exit', 'closure', 'default', 'penalty', 'fraud',
    'miss estimates', 'underperform', 'weak', 'challenging'
]


def get_bse_announcements(from_date=None, to_date=None, category=''):
    """Fetch announcements from BSE India API"""
    if not from_date:
        from_date = (datetime.now() - timedelta(days=4)).strftime('%Y%m%d')
    if not to_date:
        to_date = datetime.now().strftime('%Y%m%d')
    
    params = {
        'strCat': category,
        'strPrevDate': from_date,
        'strScrip': '',
        'strSearch': 'P',
        'strToDate': to_date,
        'strType': 'C'
    }
    
    try:
        session = requests.Session()
        # First hit the main page to get cookies
        session.get('https://www.bseindia.com/corporates/ann.html', headers=HEADERS)
        time.sleep(1)
        
        response = session.get(BSE_ANNOUNCEMENTS_URL, params=params, headers=HEADERS, timeout=30)
        response.raise_for_status()
        
        data = response.json()
        return data.get('Table', [])
    except Exception as e:
        print(f"Error fetching BSE announcements: {e}")
        return []


def get_nse_announcements(from_date=None, to_date=None):
    """Fetch announcements from NSE India API"""
    if not from_date:
        from_date = (datetime.now() - timedelta(days=4)).strftime('%d-%m-%Y')
    if not to_date:
        to_date = datetime.now().strftime('%d-%m-%Y')
    
    params = {
        'index': 'equities',
        'from_date': from_date,
        'to_date': to_date
    }
    
    try:
        session = requests.Session()
        # First hit the main page to get cookies
        session.get('https://www.nseindia.com', headers=HEADERS)
        time.sleep(2)
        
        response = session.get(NSE_ANNOUNCEMENTS_URL, params=params, headers=HEADERS, timeout=30)
        response.raise_for_status()
        
        return response.json()
    except Exception as e:
        print(f"Error fetching NSE announcements: {e}")
        return []


def extract_pdf_text(pdf_url, max_pages=3):
    """Extract text from PDF announcement"""
    try:
        response = requests.get(pdf_url, headers=HEADERS, timeout=30)
        response.raise_for_status()
        
        pdf_file = BytesIO(response.content)
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        
        text = ""
        pages_to_read = min(max_pages, len(pdf_reader.pages))
        
        for i in range(pages_to_read):
            page = pdf_reader.pages[i]
            text += page.extract_text() or ""
        
        return text.strip()
    except Exception as e:
        print(f"Error extracting PDF text from {pdf_url}: {e}")
        return ""


def categorize_announcement(subject, description=""):
    """Categorize announcement based on subject and description"""
    text = (subject + " " + description).lower()
    
    if any(kw in text for kw in ['board meeting', 'meeting of board']):
        return 'Board Meeting'
    elif any(kw in text for kw in ['financial result', 'quarterly result', 'annual result', 'q1', 'q2', 'q3', 'q4']):
        return 'Financial Results'
    elif any(kw in text for kw in ['dividend', 'interim dividend', 'final dividend']):
        return 'Dividend'
    elif any(kw in text for kw in ['agm', 'egm', 'annual general', 'extraordinary general']):
        return 'AGM/EGM'
    elif any(kw in text for kw in ['acquisition', 'acquire', 'takeover']):
        return 'Acquisition'
    elif any(kw in text for kw in ['investment', 'invest', 'stake']):
        return 'Investment'
    elif any(kw in text for kw in ['fund raising', 'qip', 'preferential', 'rights issue', 'fpo', 'ipo']):
        return 'Fund Raising'
    elif any(kw in text for kw in ['merger', 'demerger', 'amalgamation', 'scheme of arrangement']):
        return 'Merger/Demerger'
    elif any(kw in text for kw in ['director', 'appointment', 'resignation', 'cessation']):
        return 'Change in Directors'
    elif any(kw in text for kw in ['bonus', 'split', 'buyback', 'corporate action']):
        return 'Corporate Action'
    elif any(kw in text for kw in ['investor presentation', 'analyst meet', 'investor meet']):
        return 'Investor Presentation'
    elif any(kw in text for kw in ['concall', 'conference call', 'earnings call', 'transcript']):
        return 'Concall Transcript'
    elif any(kw in text for kw in ['order', 'contract', 'award', 'mandate']):
        return 'Order Win'
    elif any(kw in text for kw in ['expansion', 'capacity', 'capex', 'new plant', 'new facility']):
        return 'Expansion'
    elif any(kw in text for kw in ['rating', 'credit rating', 'upgrade', 'downgrade']):
        return 'Rating'
    else:
        return 'Others'


def extract_key_highlights(text, category):
    """Extract key highlights from announcement text using pattern matching"""
    highlights = []
    
    # Common patterns to extract
    patterns = {
        'revenue': r'revenue[:\s]+(?:rs\.?|inr|₹)?\s*([\d,\.]+)\s*(?:crore|cr|lakh|billion|million)?',
        'profit': r'(?:net\s+)?profit[:\s]+(?:rs\.?|inr|₹)?\s*([\d,\.]+)\s*(?:crore|cr|lakh|billion|million)?',
        'growth': r'(?:growth|increase|up|rose)\s+(?:of\s+)?(\d+(?:\.\d+)?)\s*%',
        'dividend': r'dividend[:\s]+(?:rs\.?|inr|₹)?\s*([\d,\.]+)\s*(?:per\s+share)?',
        'order_value': r'(?:order|contract)[:\s]+(?:worth\s+)?(?:rs\.?|inr|₹)?\s*([\d,\.]+)\s*(?:crore|cr|lakh|billion|million)?',
        'ebitda': r'ebitda[:\s]+(?:rs\.?|inr|₹)?\s*([\d,\.]+)\s*(?:crore|cr|lakh|billion|million)?',
        'margin': r'margin[:\s]+(\d+(?:\.\d+)?)\s*%',
        'eps': r'eps[:\s]+(?:rs\.?|inr|₹)?\s*([\d,\.]+)',
    }
    
    text_lower = text.lower()
    
    for metric, pattern in patterns.items():
        matches = re.findall(pattern, text_lower, re.IGNORECASE)
        if matches:
            for match in matches[:2]:  # Limit to 2 matches per metric
                highlights.append(f"{metric.upper()}: {match}")
    
    # Extract percentage changes
    pct_changes = re.findall(r'(\w+)\s+(?:increased|decreased|grew|fell|rose|dropped)\s+(?:by\s+)?(\d+(?:\.\d+)?)\s*%', text_lower)
    for metric, value in pct_changes[:3]:
        highlights.append(f"{metric.title()} change: {value}%")
    
    # If no highlights found, extract first few sentences
    if not highlights:
        sentences = text.split('.')[:3]
        highlights = [s.strip() for s in sentences if len(s.strip()) > 20]
    
    return '\n'.join(f"• {h}" for h in highlights[:8]) if highlights else "Details in PDF"


def assess_investment_implication(text, category):
    """Assess investment implication based on content"""
    text_lower = text.lower()
    
    positive_count = sum(1 for kw in POSITIVE_KEYWORDS if kw in text_lower)
    negative_count = sum(1 for kw in NEGATIVE_KEYWORDS if kw in text_lower)
    
    # Category-based assessment
    if category in ['Dividend', 'Order Win', 'Expansion']:
        positive_count += 2
    elif category in ['Fund Raising']:
        positive_count += 1  # Usually positive but dilutive
    
    if positive_count > negative_count + 2:
        return "★★★ POSITIVE"
    elif positive_count > negative_count:
        return "★★ MODERATE POSITIVE"
    elif negative_count > positive_count + 2:
        return "★ CAUTIOUS"
    elif negative_count > positive_count:
        return "★★ WATCH"
    else:
        return "★★ NEUTRAL"


def process_announcements(announcements_data):
    """Process raw announcements data into structured format"""
    processed = []
    
    for ann in announcements_data:
        try:
            # Extract fields (BSE format)
            company = ann.get('SLONGNAME', ann.get('COMPANY_NAME', 'Unknown'))
            scrip_code = ann.get('SCRIP_CD', ann.get('SYMBOL', ''))
            subject = ann.get('NEWSSUB', ann.get('SUBJECT', ''))
            news_dt = ann.get('NEWS_DT', ann.get('DATE', ''))
            attachment = ann.get('ATTACHMENTNAME', ann.get('ATTACHMENT', ''))
            
            # Parse date
            try:
                if isinstance(news_dt, str):
                    if 'T' in news_dt:
                        dt = datetime.fromisoformat(news_dt.replace('Z', '+00:00'))
                    else:
                        dt = datetime.strptime(news_dt, '%d-%b-%Y %H:%M:%S')
                else:
                    dt = datetime.now()
                date_str = dt.strftime('%d %B %Y')
                time_str = dt.strftime('%I:%M %p')
            except:
                date_str = str(news_dt)
                time_str = ''
            
            # Build PDF URL
            if attachment:
                pdf_url = f"{BSE_PDF_BASE_URL}/{attachment}"
            else:
                pdf_url = ''
            
            # Categorize
            category = categorize_announcement(subject)
            
            # Extract highlights (from subject for now, can enhance with PDF parsing)
            highlights = extract_key_highlights(subject, category)
            
            # Assess investment implication
            implication = assess_investment_implication(subject + " " + highlights, category)
            
            processed.append({
                'Company': company,
                'Scrip Code': scrip_code,
                'Category': category,
                'Subject': subject,
                'Date': date_str,
                'Time': time_str,
                'Key Highlights': highlights,
                'Investment Implication': implication,
                'PDF Link': pdf_url
            })
            
        except Exception as e:
            print(f"Error processing announcement: {e}")
            continue
    
    return processed


def create_excel_report(data, output_file='India_Corporate_Announcements.xlsx'):
    """Create formatted Excel report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Announcements"
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Color fills for implications
    positive_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    moderate_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    cautious_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Headers
    headers = ['Company', 'Scrip Code', 'Category', 'Subject', 'Date', 'Time', 
               'Key Highlights', 'Investment Implication', 'PDF Link']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_num, item in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=item['Company']).alignment = cell_alignment
        ws.cell(row=row_num, column=2, value=item['Scrip Code']).alignment = cell_alignment
        ws.cell(row=row_num, column=3, value=item['Category']).alignment = cell_alignment
        ws.cell(row=row_num, column=4, value=item['Subject']).alignment = cell_alignment
        ws.cell(row=row_num, column=5, value=item['Date']).alignment = cell_alignment
        ws.cell(row=row_num, column=6, value=item['Time']).alignment = cell_alignment
        ws.cell(row=row_num, column=7, value=item['Key Highlights']).alignment = cell_alignment
        ws.cell(row=row_num, column=8, value=item['Investment Implication']).alignment = cell_alignment
        ws.cell(row=row_num, column=9, value=item['PDF Link']).alignment = cell_alignment
        
        # Apply borders
        for col in range(1, 10):
            ws.cell(row=row_num, column=col).border = thin_border
        
        # Color code investment implication
        impl = item['Investment Implication']
        impl_cell = ws.cell(row=row_num, column=8)
        if '★★★' in impl:
            impl_cell.fill = positive_fill
        elif 'CAUTIOUS' in impl:
            impl_cell.fill = cautious_fill
        elif '★★' in impl:
            impl_cell.fill = moderate_fill
    
    # Column widths
    column_widths = [30, 12, 18, 50, 15, 12, 45, 25, 60]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Row heights
    ws.row_dimensions[1].height = 25
    for row_num in range(2, len(data) + 2):
        ws.row_dimensions[row_num].height = 80
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:I{len(data) + 1}"
    
    wb.save(output_file)
    print(f"Excel report saved: {output_file}")
    return output_file


def create_summary_sheet(data, output_file):
    """Add a summary sheet to the Excel file"""
    wb = load_workbook(output_file)
    
    # Create summary sheet
    if 'Summary' in wb.sheetnames:
        del wb['Summary']
    
    ws = wb.create_sheet('Summary', 0)
    
    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    
    # Title
    ws['A1'] = 'Indian Corporate Announcements Summary'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f'Generated: {datetime.now().strftime("%d %B %Y %I:%M %p")}'
    
    # Category breakdown
    ws['A4'] = 'Announcements by Category'
    ws['A4'].font = header_font
    
    category_counts = {}
    for item in data:
        cat = item['Category']
        category_counts[cat] = category_counts.get(cat, 0) + 1
    
    row = 5
    for cat, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True):
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=count)
        row += 1
    
    # Investment implications breakdown
    row += 2
    ws.cell(row=row, column=1, value='Investment Implications Breakdown')
    ws.cell(row=row, column=1).font = header_font
    row += 1
    
    impl_counts = {}
    for item in data:
        impl = item['Investment Implication'].split()[0]  # Get star rating
        impl_counts[impl] = impl_counts.get(impl, 0) + 1
    
    for impl, count in sorted(impl_counts.items(), reverse=True):
        ws.cell(row=row, column=1, value=impl)
        ws.cell(row=row, column=2, value=count)
        row += 1
    
    # Top companies
    row += 2
    ws.cell(row=row, column=1, value='Companies with Most Announcements')
    ws.cell(row=row, column=1).font = header_font
    row += 1
    
    company_counts = {}
    for item in data:
        comp = item['Company']
        company_counts[comp] = company_counts.get(comp, 0) + 1
    
    for comp, count in sorted(company_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
        ws.cell(row=row, column=1, value=comp)
        ws.cell(row=row, column=2, value=count)
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    
    wb.save(output_file)
    print("Summary sheet added")


def main():
    """Main function to run the scraper"""
    print("=" * 60)
    print("Indian Corporate Announcements Scraper")
    print("=" * 60)
    
    # Calculate date range (last 4 days)
    to_date = datetime.now()
    from_date = to_date - timedelta(days=4)
    
    print(f"\nFetching announcements from {from_date.strftime('%d-%b-%Y')} to {to_date.strftime('%d-%b-%Y')}")
    
    # Fetch BSE announcements
    print("\nFetching BSE announcements...")
    bse_data = get_bse_announcements(
        from_date=from_date.strftime('%Y%m%d'),
        to_date=to_date.strftime('%Y%m%d')
    )
    print(f"Found {len(bse_data)} BSE announcements")
    
    # Process announcements
    print("\nProcessing announcements...")
    processed_data = process_announcements(bse_data)
    print(f"Processed {len(processed_data)} announcements")
    
    if processed_data:
        # Sort by date (most recent first)
        processed_data.sort(key=lambda x: x['Date'], reverse=True)
        
        # Create output directory
        output_dir = os.environ.get('OUTPUT_DIR', '.')
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate filename with date
        output_file = os.path.join(
            output_dir, 
            f"India_Corporate_Announcements_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        # Create Excel report
        print("\nCreating Excel report...")
        create_excel_report(processed_data, output_file)
        
        # Add summary sheet
        create_summary_sheet(processed_data, output_file)
        
        print(f"\n✅ Report generated successfully: {output_file}")
        print(f"Total announcements: {len(processed_data)}")
        
        # Print category summary
        print("\nCategory Summary:")
        category_counts = {}
        for item in processed_data:
            cat = item['Category']
            category_counts[cat] = category_counts.get(cat, 0) + 1
        
        for cat, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True):
            print(f"  {cat}: {count}")
    else:
        print("\n⚠️ No announcements found for the specified date range")
    
    print("\n" + "=" * 60)


if __name__ == "__main__":
    main()
