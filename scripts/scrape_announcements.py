#!/usr/bin/env python3
"""
Indian Corporate Announcements Scraper
Fetches announcements from BSE India, updates Google Sheets, and sends email
"""

import requests
import pandas as pd
from datetime import datetime, timedelta
import json
import re
import os
import time
import base64
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# Google Sheets imports
try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    GOOGLE_AVAILABLE = True
except ImportError:
    GOOGLE_AVAILABLE = False
    print("Google API libraries not available")

# Configuration
BSE_ANNOUNCEMENTS_URL = "https://api.bseindia.com/BseIndiaAPI/api/AnnGetData/w"
BSE_PDF_BASE_URL = "https://www.bseindia.com/xml-data/corpfiling/AttachLive"

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'en-US,en;q=0.9',
    'Referer': 'https://www.bseindia.com/',
    'Origin': 'https://www.bseindia.com'
}

# Keywords for investment implications
POSITIVE_KEYWORDS = [
    'profit increase', 'profit up', 'revenue growth', 'dividend', 'bonus',
    'acquisition', 'expansion', 'new order', 'contract win', 'upgrade',
    'record', 'highest ever', 'beat estimates', 'outperform', 'growth',
    'investment', 'capex', 'expansion plan', 'new plant', 'capacity addition'
]

NEGATIVE_KEYWORDS = [
    'profit decline', 'profit down', 'revenue decline', 'loss', 'downgrade',
    'resign', 'exit', 'closure', 'default', 'penalty', 'fraud',
    'miss estimates', 'underperform', 'weak', 'challenging'
]


def get_bse_announcements_by_date(target_date):
    """Fetch announcements from BSE India API for a specific date"""
    date_str = target_date.strftime('%Y%m%d')
    
    params = {
        'strCat': '-1',
        'strPrevDate': date_str,
        'strScrip': '',
        'strSearch': 'P',
        'strToDate': date_str,
        'strType': 'C'
    }
    
    try:
        session = requests.Session()
        session.get('https://www.bseindia.com/corporates/ann.html', headers=HEADERS, timeout=10)
        time.sleep(0.5)
        
        response = session.get(BSE_ANNOUNCEMENTS_URL, params=params, headers=HEADERS, timeout=30)
        response.raise_for_status()
        
        data = response.json()
        table_data = data.get('Table', [])
        print(f"  {target_date.strftime('%d-%b-%Y')}: Found {len(table_data)} announcements")
        return table_data
    except Exception as e:
        print(f"  Error fetching {target_date.strftime('%d-%b-%Y')}: {e}")
        return []


def get_bse_announcements_multi_day(days_back=4):
    """Fetch announcements for multiple days"""
    all_announcements = []
    
    for i in range(days_back):
        target_date = datetime.now() - timedelta(days=i)
        daily_data = get_bse_announcements_by_date(target_date)
        all_announcements.extend(daily_data)
        time.sleep(1)  # Be nice to the API
    
    return all_announcements


def categorize_announcement(subject, description=""):
    """Categorize announcement based on subject and description"""
    text = (subject + " " + description).lower()
    
    if any(kw in text for kw in ['board meeting', 'meeting of board']):
        return 'Board Meeting'
    elif any(kw in text for kw in ['financial result', 'quarterly result', 'annual result', 'q1', 'q2', 'q3', 'q4']):
        return 'Financial Results'
    elif any(kw in text for kw in ['dividend', 'interim dividend', 'final dividend']):
        return 'Dividend'
    elif any(kw in text for kw in ['agm', 'egm', 'annual general', 'extraordinary general']):
        return 'AGM/EGM'
    elif any(kw in text for kw in ['acquisition', 'acquire', 'takeover']):
        return 'Acquisition'
    elif any(kw in text for kw in ['investor presentation', 'analyst meet', 'investor meet']):
        return 'Investor Presentation'
    elif any(kw in text for kw in ['fund raising', 'qip', 'preferential', 'rights issue', 'fpo']):
        return 'Fund Raising'
    elif any(kw in text for kw in ['merger', 'demerger', 'amalgamation', 'scheme of arrangement']):
        return 'Merger/Demerger'
    elif any(kw in text for kw in ['director', 'appointment', 'resignation', 'cessation']):
        return 'Change in Directors'
    elif any(kw in text for kw in ['bonus', 'split', 'buyback', 'corporate action']):
        return 'Corporate Action'
    elif any(kw in text for kw in ['concall', 'conference call', 'earnings call', 'transcript']):
        return 'Concall Transcript'
    elif any(kw in text for kw in ['order', 'contract', 'award', 'mandate']):
        return 'Order Win'
    elif any(kw in text for kw in ['expansion', 'capacity', 'capex', 'new plant', 'new facility']):
        return 'Expansion'
    elif any(kw in text for kw in ['rating', 'credit rating', 'upgrade', 'downgrade']):
        return 'Rating'
    else:
        return 'Others'


def extract_key_highlights(text, category):
    """Extract key highlights from announcement text"""
    highlights = []
    
    patterns = {
        'revenue': r'revenue[:\s]+(?:rs\.?|inr|₹)?\s*([\d,\.]+)\s*(?:crore|cr|lakh|billion|million)?',
        'profit': r'(?:net\s+)?profit[:\s]+(?:rs\.?|inr|₹)?\s*([\d,\.]+)\s*(?:crore|cr|lakh|billion|million)?',
        'growth': r'(?:growth|increase|up|rose)\s+(?:of\s+)?(\d+(?:\.\d+)?)\s*%',
        'dividend': r'dividend[:\s]+(?:rs\.?|inr|₹)?\s*([\d,\.]+)\s*(?:per\s+share)?',
    }
    
    text_lower = text.lower()
    
    for metric, pattern in patterns.items():
        matches = re.findall(pattern, text_lower, re.IGNORECASE)
        if matches:
            for match in matches[:2]:
                highlights.append(f"{metric.upper()}: {match}")
    
    pct_changes = re.findall(r'(\w+)\s+(?:increased|decreased|grew|fell|rose|dropped)\s+(?:by\s+)?(\d+(?:\.\d+)?)\s*%', text_lower)
    for metric, value in pct_changes[:3]:
        highlights.append(f"{metric.title()} change: {value}%")
    
    if not highlights:
        sentences = text.split('.')[:2]
        highlights = [s.strip()[:100] for s in sentences if len(s.strip()) > 20]
    
    return ' | '.join(highlights[:4]) if highlights else "See PDF for details"


def assess_investment_implication(text, category):
    """Assess investment implication based on content"""
    text_lower = text.lower()
    
    positive_count = sum(1 for kw in POSITIVE_KEYWORDS if kw in text_lower)
    negative_count = sum(1 for kw in NEGATIVE_KEYWORDS if kw in text_lower)
    
    if category in ['Dividend', 'Order Win', 'Expansion']:
        positive_count += 2
    elif category in ['Fund Raising']:
        positive_count += 1
    
    if positive_count > negative_count + 2:
        return "★★★ POSITIVE"
    elif positive_count > negative_count:
        return "★★ MODERATE"
    elif negative_count > positive_count + 2:
        return "★ CAUTIOUS"
    elif negative_count > positive_count:
        return "★★ WATCH"
    else:
        return "★★ NEUTRAL"


def process_announcements(announcements_data):
    """Process raw announcements data into structured format"""
    processed = []
    seen = set()
    
    for ann in announcements_data:
        try:
            company = ann.get('SLONGNAME', ann.get('COMPANY_NAME', 'Unknown'))
            scrip_code = str(ann.get('SCRIP_CD', ann.get('SYMBOL', '')))
            subject = ann.get('NEWSSUB', ann.get('SUBJECT', ''))
            news_dt = ann.get('NEWS_DT', ann.get('DATE', ''))
            attachment = ann.get('ATTACHMENTNAME', ann.get('ATTACHMENT', ''))
            
            # Create unique key to avoid duplicates
            unique_key = f"{scrip_code}_{subject[:50]}_{news_dt}"
            if unique_key in seen:
                continue
            seen.add(unique_key)
            
            # Parse date
            try:
                if isinstance(news_dt, str):
                    if 'T' in news_dt:
                        dt = datetime.fromisoformat(news_dt.replace('Z', '+00:00'))
                    else:
                        dt = datetime.strptime(news_dt, '%d-%b-%Y %H:%M:%S')
                else:
                    dt = datetime.now()
                date_str = dt.strftime('%d %b %Y')
                time_str = dt.strftime('%I:%M %p')
            except:
                date_str = str(news_dt)[:11] if news_dt else ''
                time_str = ''
            
            # Build PDF URL
            pdf_url = f"{BSE_PDF_BASE_URL}/{attachment}" if attachment else ''
            
            # Categorize
            category = categorize_announcement(subject)
            
            # Extract highlights
            highlights = extract_key_highlights(subject, category)
            
            # Assess investment implication
            implication = assess_investment_implication(subject + " " + highlights, category)
            
            processed.append({
                'Company': company,
                'Scrip Code': scrip_code,
                'Category': category,
                'Subject': subject[:200],
                'Date': date_str,
                'Time': time_str,
                'Key Highlights': highlights,
                'Investment Implication': implication,
                'PDF Link': pdf_url
            })
            
        except Exception as e:
            print(f"Error processing announcement: {e}")
            continue
    
    return processed


def create_excel_report(data, output_file='India_Corporate_Announcements.xlsx'):
    """Create formatted Excel report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Announcements"
    
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    positive_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    moderate_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    cautious_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    headers = ['Company', 'Scrip Code', 'Category', 'Subject', 'Date', 'Time', 
               'Key Highlights', 'Investment Implication', 'PDF Link']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_num, item in enumerate(data, 2):
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=item.get(key, ''))
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        impl = item.get('Investment Implication', '')
        impl_cell = ws.cell(row=row_num, column=8)
        if '★★★' in impl:
            impl_cell.fill = positive_fill
        elif 'CAUTIOUS' in impl:
            impl_cell.fill = cautious_fill
        elif '★★' in impl:
            impl_cell.fill = moderate_fill
    
    column_widths = [30, 12, 18, 50, 12, 10, 40, 20, 50]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:I{len(data) + 1}"
    
    wb.save(output_file)
    print(f"Excel report saved: {output_file}")
    return output_file


def update_google_sheet(data):
    """Update Google Sheet with announcement data"""
    if not GOOGLE_AVAILABLE:
        print("Google API not available, skipping sheet update")
        return False
    
    credentials_json = os.environ.get('GOOGLE_CREDENTIALS')
    sheet_id = os.environ.get('GOOGLE_SHEET_ID')
    
    if not credentials_json or not sheet_id:
        print("Google credentials or sheet ID not configured")
        return False
    
    try:
        # Parse credentials
        creds_dict = json.loads(credentials_json)
        credentials = service_account.Credentials.from_service_account_info(
            creds_dict,
            scopes=['https://www.googleapis.com/auth/spreadsheets']
        )
        
        # Build service
        service = build('sheets', 'v4', credentials=credentials)
        sheet = service.spreadsheets()
        
        # Prepare data
        headers = ['Company', 'Scrip Code', 'Category', 'Subject', 'Date', 'Time', 
                   'Key Highlights', 'Investment Implication', 'PDF Link']
        
        rows = [headers]
        for item in data:
            row = [item.get(h, '') for h in headers]
            rows.append(row)
        
        # Clear existing data
        sheet.values().clear(
            spreadsheetId=sheet_id,
            range='Sheet1!A:I'
        ).execute()
        
        # Write new data
        sheet.values().update(
            spreadsheetId=sheet_id,
            range='Sheet1!A1',
            valueInputOption='RAW',
            body={'values': rows}
        ).execute()
        
        # Add timestamp
        timestamp = datetime.now().strftime('%d %b %Y %I:%M %p')
        sheet.values().update(
            spreadsheetId=sheet_id,
            range='Sheet1!K1',
            valueInputOption='RAW',
            body={'values': [[f'Last Updated: {timestamp}']]}
        ).execute()
        
        print(f"Google Sheet updated with {len(data)} announcements")
        return True
        
    except Exception as e:
        print(f"Error updating Google Sheet: {e}")
        return False


def send_email_report(output_file, announcement_count):
    """Send email with Excel report attached"""
    email_address = os.environ.get('EMAIL_ADDRESS')
    email_password = os.environ.get('EMAIL_PASSWORD')
    notify_email = os.environ.get('NOTIFY_EMAIL', email_address)
    
    if not email_address or not email_password:
        print("Email credentials not configured, skipping email")
        return False
    
    try:
        msg = MIMEMultipart()
        msg['From'] = email_address
        msg['To'] = notify_email
        msg['Subject'] = f"📊 India Corporate Announcements - {datetime.now().strftime('%d %b %Y')} ({announcement_count} updates)"
        
        # Email body
        body = f"""
Hello,

Your daily India Corporate Announcements report is ready!

📊 Summary:
- Total Announcements: {announcement_count}
- Report Date: {datetime.now().strftime('%d %B %Y')}
- Time: {datetime.now().strftime('%I:%M %p IST')}

The Excel report is attached. You can also view the live data in your Google Sheet.

Google Sheet: https://docs.google.com/spreadsheets/d/{os.environ.get('GOOGLE_SHEET_ID', '')}/edit

---
This is an automated report from your India Corporate Announcements Tracker.
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach Excel file
        with open(output_file, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {os.path.basename(output_file)}'
            )
            msg.attach(part)
        
        # Send email
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(email_address, email_password)
        server.send_message(msg)
        server.quit()
        
        print(f"Email sent to {notify_email}")
        return True
        
    except Exception as e:
        print(f"Error sending email: {e}")
        return False


def main():
    """Main function to run the scraper"""
    print("=" * 60)
    print("Indian Corporate Announcements Scraper")
    print("=" * 60)
    
    # Get days_back from environment or default to 4
    days_back = int(os.environ.get('DAYS_BACK', 4))
    
    print(f"\nFetching announcements for last {days_back} days...")
    
    # Fetch announcements for multiple days
    bse_data = get_bse_announcements_multi_day(days_back)
    print(f"\nTotal raw announcements: {len(bse_data)}")
    
    # Process announcements
    print("\nProcessing announcements...")
    processed_data = process_announcements(bse_data)
    print(f"Processed {len(processed_data)} unique announcements")
    
    if processed_data:
        # Sort by date (most recent first)
        processed_data.sort(key=lambda x: x['Date'], reverse=True)
        
        # Create output directory
        output_dir = os.environ.get('OUTPUT_DIR', '.')
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate filename
        output_file = os.path.join(
            output_dir, 
            f"India_Corporate_Announcements_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        # Create Excel report
        print("\nCreating Excel report...")
        create_excel_report(processed_data, output_file)
        
        # Update Google Sheet
        print("\nUpdating Google Sheet...")
        update_google_sheet(processed_data)
        
        # Send email
        print("\nSending email report...")
        send_email_report(output_file, len(processed_data))
        
        print(f"\n✅ Report generated successfully!")
        print(f"Total announcements: {len(processed_data)}")
        
        # Print category summary
        print("\nCategory Summary:")
        category_counts = {}
        for item in processed_data:
            cat = item['Category']
            category_counts[cat] = category_counts.get(cat, 0) + 1
        
        for cat, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True):
            print(f"  {cat}: {count}")
    else:
        print("\n⚠️ No announcements found")
    
    print("\n" + "=" * 60)


if __name__ == "__main__":
    main()
